from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from database.models import DepartmentModel, DocumentModel, PositionModel, UserDocumentModel, UserModel, VersionModel, ApproverModel

# --- Лист «Отделы»: [0] Номер отдела, [1] Наименование отдела ---
COL_DEP_ID = 0
COL_DEP_TITLE = 1

# --- Лист «Должности»: [0] Сокр. наименование, [1] Полное наименование ---
COL_POS_SHORT = 0
COL_POS_FULL = 1

# --- «Руководители» / «Сотрудники» (одинаковые индексы для общих полей) ---
# 0 Таб.№ → users.id (VARCHAR)
# 1 Фамилия → surname
# 2 Имя → first_name
# 3 Отчество → second_name
# 4 Дата рождения → birth_date
# 5 Пол → sex
# 6 Логин → login
# 7 Отдел / Номер отдела → department_id (строка, FK departments.id)
# 8 Должность → position_id (строка, FK positions.short_name)
# 9 только «Сотрудники»: Таб.№ руководителя → manager_id (VARCHAR, FK users.id)
U_COL_ID = 0
U_COL_SURNAME = 1
U_COL_FIRST = 2
U_COL_PATRONYMIC = 3
U_COL_BIRTH = 4
U_COL_SEX = 5
U_COL_LOGIN = 6
U_COL_DEPT = 7
U_COL_POSITION = 8
U_COL_MANAGER = 9

MGR_MIN_COLS = 9
EMP_MIN_COLS = 10


class UserImportService:
    def __init__(self, db: Session) -> None:
        self._db = db

    @staticmethod
    def _cell(values: tuple[Any, ...] | list[Any] | None, index: int) -> Any:
        if not values or index >= len(values):
            return None
        return values[index]

    @staticmethod
    def _normalize_str(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _department_id_str(value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        if isinstance(value, int):
            return str(value)
        return UserImportService._normalize_str(value)

    @staticmethod
    def _parse_sex(value: Any) -> bool | None:
        raw = UserImportService._normalize_str(value)
        if raw is None:
            return None
        normalized = raw.lower().rstrip(".")
        if normalized in {"м", "муж", "мужчина", "male", "m", "1", "true"}:
            return True
        if normalized in {"ж", "жен", "женщина", "female", "f", "0", "false"}:
            return False
        return None

    @staticmethod
    def _parse_birth_date(value: Any) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        raw = UserImportService._normalize_str(value)
        if raw is None:
            return None
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_int(value: Any) -> int | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float) and value.is_integer():
            return int(value)
        raw = UserImportService._normalize_str(value)
        if raw is None:
            return None
        try:
            return int(raw)
        except ValueError:
            try:
                f = float(raw.replace(",", "."))
                return int(f) if f.is_integer() else None
            except ValueError:
                return None

    def _upsert_department(self, department_id: str | None, department_title: str | None) -> None:
        if department_id is None and department_title is None:
            return
        dep = None
        if department_id is not None:
            dep = self._db.get(DepartmentModel, department_id)
        if dep is None and department_title is not None:
            dep = self._db.execute(
                select(DepartmentModel).where(DepartmentModel.department_title == department_title)
            ).scalars().first()
        if dep is None:
            did = department_id or department_title
            dtitle = department_title or department_id
            self._db.add(DepartmentModel(id=did, department_title=dtitle))
            self._db.flush()
            return
        if department_title is not None and dep.department_title != department_title:
            dep.department_title = department_title

    def _upsert_position_row(self, short_name: str | None, full_name: str | None) -> None:
        if short_name is None and full_name is None:
            return
        short = short_name or full_name
        full = full_name or short_name or short
        pos = self._db.execute(select(PositionModel).where(PositionModel.short_name == short)).scalars().first()
        if pos is None:
            self._db.add(PositionModel(short_name=short, full_name=full))
            self._db.flush()
            return
        if full_name is not None and pos.full_name != full:
            pos.full_name = full
        if short_name is not None and pos.short_name != short:
            pos.short_name = short

    def _import_departments(self, worksheet: Any, result: dict[str, Any]) -> None:
        for row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not values or all(v is None for v in values):
                continue
            dep_id = self._department_id_str(self._cell(values, COL_DEP_ID))
            dep_title = self._normalize_str(self._cell(values, COL_DEP_TITLE))
            if dep_id is None and dep_title is None:
                continue
            result["total_rows"] += 1
            try:
                self._upsert_department(dep_id, dep_title)
            except Exception as exc:
                result["errors"].append({"sheet": "Отделы", "row": row_index, "message": str(exc)})

    def _import_positions(self, worksheet: Any, result: dict[str, Any]) -> None:
        for row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not values or all(v is None for v in values):
                continue
            short_n = self._normalize_str(self._cell(values, COL_POS_SHORT))
            full_n = self._normalize_str(self._cell(values, COL_POS_FULL))
            if short_n is None and full_n is None:
                continue
            result["total_rows"] += 1
            try:
                self._upsert_position_row(short_n, full_n)
            except Exception as exc:
                result["errors"].append({"sheet": "Должности", "row": row_index, "message": str(exc)})

    def import_users_from_xlsx(self, file_content: bytes, skip_user_ids: frozenset[str] | None = None) -> dict[str, Any]:
        workbook = load_workbook(filename=BytesIO(file_content), data_only=True)
        names = workbook.sheetnames

        if "Руководители" not in names and "Сотрудники" not in names:
            raise ValueError("В файле не найдены листы 'Руководители' или 'Сотрудники'")

        result: dict[str, Any] = {
            "total_rows": 0,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [],
        }

        if "Отделы" in names:
            self._import_departments(workbook["Отделы"], result)
        if "Должности" in names:
            self._import_positions(workbook["Должности"], result)

        for sheet_name in ("Руководители", "Сотрудники"):
            if sheet_name not in names:
                continue
            worksheet = workbook[sheet_name]
            is_managers = sheet_name == "Руководители"
            min_cols = MGR_MIN_COLS if is_managers else EMP_MIN_COLS

            for row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not values or all(v is None for v in values):
                    continue
                if len(values) < min_cols:
                    result["skipped"] += 1
                    result["errors"].append(
                        {"sheet": sheet_name, "row": row_index, "message": f"Нужно минимум {min_cols} колонок"}
                    )
                    continue

                tab_id = self._department_id_str(self._cell(values, U_COL_ID))
                login = self._normalize_str(self._cell(values, U_COL_LOGIN))
                if not tab_id or not login:
                    result["skipped"] += 1
                    result["errors"].append(
                        {"sheet": sheet_name, "row": row_index, "message": "Пустой табельный номер или логин"}
                    )
                    continue

                if skip_user_ids is not None and tab_id in skip_user_ids:
                    result["skipped"] += 1
                    continue

                result["total_rows"] += 1

                try:
                    user = self._db.get(UserModel, tab_id)
                    if user is None:
                        user = UserModel(id=tab_id, login=login)
                        self._db.add(user)
                        result["created"] += 1
                    else:
                        user.login = login
                        result["updated"] += 1

                    user.surname = self._normalize_str(self._cell(values, U_COL_SURNAME))
                    user.first_name = self._normalize_str(self._cell(values, U_COL_FIRST))
                    user.second_name = self._normalize_str(self._cell(values, U_COL_PATRONYMIC))
                    user.birth_date = self._parse_birth_date(self._cell(values, U_COL_BIRTH))
                    user.sex = self._parse_sex(self._cell(values, U_COL_SEX))
                    user.department_id = self._department_id_str(self._cell(values, U_COL_DEPT))
                    user.position_id = self._normalize_str(self._cell(values, U_COL_POSITION))

                    if is_managers:
                        user.manager_id = None
                    else:
                        mid = self._department_id_str(self._cell(values, U_COL_MANAGER))
                        user.manager_id = mid if mid and mid != tab_id else None

                except Exception as exc:
                    result["skipped"] += 1
                    result["errors"].append({"sheet": sheet_name, "row": row_index, "message": str(exc)})

        self._db.commit()
        return result

    def replace_all_users_from_xlsx(self, file_content: bytes) -> dict[str, Any]:
        protected_ids = frozenset(
            u.id for u in self._db.query(UserModel).filter(UserModel.is_superuser.is_(True)).all()
        )
        
        # Удаляем в правильном порядке (от зависимых к главным)
        self._db.query(ApproverModel).delete(synchronize_session=False)      # 1. Удаляем аппруверов
        self._db.query(UserDocumentModel).delete(synchronize_session=False)  # 2. Удаляем связи пользователь-документ
        self._db.query(VersionModel).delete(synchronize_session=False)       # 3. Удаляем версии
        self._db.query(DocumentModel).delete(synchronize_session=False)      # 4. Удаляем документы
        self._db.query(UserModel).filter(UserModel.is_superuser.is_(False)).delete(synchronize_session=False)  # 5. Удаляем обычных пользователей
        
        self._db.commit()
        return self.import_users_from_xlsx(file_content, skip_user_ids=protected_ids)
